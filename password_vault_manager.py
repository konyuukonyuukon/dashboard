"""
password_vault_manager.py
パスワード管理ファイル（password_vault.xlsx）をPythonで操作するツール

使い方:
  python password_vault_manager.py              # 対話メニュー
  python password_vault_manager.py --list       # 全エントリ一覧表示
  python password_vault_manager.py --search Amazon  # キーワード検索
  python password_vault_manager.py --file 別ファイル.xlsx  # ファイル指定
"""

import openpyxl
import argparse
import getpass
import sys
from pathlib import Path

VAULT_FILE = "password_vault.xlsx"

# データはB列(index=1)から始まる
DATA_COL_OFFSET = 1

CATEGORIES_PASSWORD = ["金融・銀行", "SNS・メール", "ショッピング", "医療・行政", "その他サービス"]
CATEGORIES_PIN      = ["【夫】", "【妻】"]


# ─────────────────────────────────────────────
# ユーティリティ
# ─────────────────────────────────────────────

def load_wb(filepath: str):
    p = Path(filepath)
    if not p.exists():
        print(f"❌ ファイルが見つかりません: {filepath}")
        sys.exit(1)
    return openpyxl.load_workbook(filepath)


def get(row, col_offset):
    """row(tuple)からB列起点でインデックスを取得"""
    idx = col_offset + DATA_COL_OFFSET
    return row[idx] if idx < len(row) else None


def is_skip_row(first_val: str) -> bool:
    skip_prefixes = ("🔐", "💳", "📱", "🛍️", "🏥", "🔧", "⚠️", "💡", "🔢",
                     "サービス名", "カード/機器名", "【夫】", "【妻】")
    return any(first_val.startswith(p) for p in skip_prefixes)


# ─────────────────────────────────────────────
# 読み込み
# ─────────────────────────────────────────────

def read_password_sheet(ws) -> list[dict]:
    entries = []
    current_category = ""
    for row in ws.iter_rows(values_only=True):
        first = str(row[DATA_COL_OFFSET] or "")
        if not first:
            continue
        # カテゴリ行
        if any(kw in first for kw in CATEGORIES_PASSWORD):
            current_category = first.lstrip("💳📱🛍️🏥🔧 ").strip()
            continue
        # スキップ行
        if is_skip_row(first):
            continue
        entries.append({
            "カテゴリ":          current_category,
            "サービス名":        row[1],
            "ログインID":        row[2] if len(row) > 2 else None,
            "パスワード":        row[3] if len(row) > 3 else None,
            "PIN/暗証番号":      row[4] if len(row) > 4 else None,
            "2FA":              row[5] if len(row) > 5 else None,
            "バックアップコード": row[6] if len(row) > 6 else None,
            "備考":             row[7] if len(row) > 7 else None,
        })
    return entries


def read_pin_sheet(ws) -> list[dict]:
    entries = []
    current_owner = ""
    for row in ws.iter_rows(values_only=True):
        first = str(row[DATA_COL_OFFSET] or "")
        if not first:
            continue
        if "【夫】" in first:
            current_owner = "夫"; continue
        if "【妻】" in first:
            current_owner = "妻"; continue
        if is_skip_row(first):
            continue
        entries.append({
            "所有者":      current_owner,
            "カード/機器名": row[1],
            "PIN/暗証番号": row[2] if len(row) > 2 else None,
            "桁数":        row[3] if len(row) > 3 else None,
            "更新日":      row[4] if len(row) > 4 else None,
            "備考":        row[5] if len(row) > 5 else None,
        })
    return entries


def load_all(filepath: str = VAULT_FILE) -> dict:
    wb = load_wb(filepath)
    result = {}
    for name in wb.sheetnames:
        ws = wb[name]
        if "PIN" in name or "暗証番号" in name:
            result[name] = read_pin_sheet(ws)
        elif "パスワード" in name:
            result[name] = read_password_sheet(ws)
    wb.close()
    return result


# ─────────────────────────────────────────────
# 書き込み（追加・削除）
# ─────────────────────────────────────────────

def add_entry(filepath: str, sheet_name: str, category_keyword: str, data: dict) -> bool:
    """カテゴリ末尾に1行追加"""
    wb = load_wb(filepath)
    if sheet_name not in wb.sheetnames:
        print(f"❌ シート '{sheet_name}' が見つかりません。")
        wb.close(); return False

    ws = wb[sheet_name]
    is_pin = "PIN" in sheet_name or "暗証番号" in sheet_name

    # カテゴリ行とその次のカテゴリ行を探す
    cat_row = None
    next_cat_row = None
    found = False

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        first = str(row[DATA_COL_OFFSET] or "")
        if not found:
            if category_keyword in first:
                cat_row = i; found = True
        else:
            if is_pin:
                if "【" in first and category_keyword not in first:
                    next_cat_row = i; break
            else:
                if any(kw in first for kw in CATEGORIES_PASSWORD) and category_keyword not in first:
                    next_cat_row = i; break

    if cat_row is None:
        print(f"❌ カテゴリ '{category_keyword}' が見つかりません。")
        wb.close(); return False

    # 挿入行を決定：次カテゴリの直前の最終データ行の次
    boundary = next_cat_row if next_cat_row else ws.max_row + 2
    insert_row = cat_row + 1
    for r in range(boundary - 1, cat_row, -1):
        if ws.cell(row=r, column=DATA_COL_OFFSET + 1).value not in (None, ""):
            insert_row = r + 1
            break

    ws.insert_rows(insert_row)

    if is_pin:
        vals = [data.get("カード/機器名",""), data.get("PIN/暗証番号",""),
                data.get("桁数",""), data.get("更新日",""), data.get("備考","")]
    else:
        vals = [data.get("サービス名",""),  data.get("ログインID",""),
                data.get("パスワード",""),   data.get("PIN/暗証番号",""),
                data.get("2FA",""),          data.get("バックアップコード",""),
                data.get("備考","")]

    for j, v in enumerate(vals):
        ws.cell(row=insert_row, column=DATA_COL_OFFSET + 1 + j, value=v)

    wb.save(filepath)
    wb.close()
    print(f"✅ 追加しました（シート: {sheet_name}、行: {insert_row}）")
    return True


def delete_entry(filepath: str, sheet_name: str, name: str) -> bool:
    """サービス名/カード名が一致する行を削除"""
    wb = load_wb(filepath)
    if sheet_name not in wb.sheetnames:
        print(f"❌ シート '{sheet_name}' が見つかりません。")
        wb.close(); return False

    ws = wb[sheet_name]
    for row in ws.iter_rows():
        if str(row[DATA_COL_OFFSET].value or "") == name:
            ws.delete_rows(row[DATA_COL_OFFSET].row)
            wb.save(filepath)
            wb.close()
            print(f"✅ '{name}' を削除しました。")
            return True

    print(f"⚠️  '{name}' は見つかりませんでした。")
    wb.close()
    return False


# ─────────────────────────────────────────────
# 検索・表示
# ─────────────────────────────────────────────

def search_entries(data: dict, keyword: str) -> list[tuple]:
    results = []
    for sheet_name, entries in data.items():
        for entry in entries:
            if any(keyword in str(v or "") for v in entry.values()):
                results.append((sheet_name, entry))
    return results


def print_entry(sheet_name: str, entry: dict):
    label = sheet_name.replace("のパスワード", "").strip()
    print(f"\n  📌 [{label}]")
    for k, v in entry.items():
        if v not in (None, ""):
            label_str = "パスワード" if k == "パスワード" else k
            # パスワード・PINはマスク表示（表示切替可）
            if k in ("パスワード", "PIN/暗証番号") and v:
                display = "****"
            else:
                display = v
            print(f"     {label_str}: {display}")


def print_entry_full(sheet_name: str, entry: dict):
    """マスクなしで表示（検索結果詳細確認用）"""
    label = sheet_name.replace("のパスワード", "").strip()
    print(f"\n  📌 [{label}]")
    for k, v in entry.items():
        if v not in (None, ""):
            print(f"     {k}: {v}")


def list_all(data: dict, masked: bool = True):
    for sheet_name, entries in data.items():
        print(f"\n{'='*50}")
        print(f"  {sheet_name}")
        print(f"{'='*50}")
        for entry in entries:
            if masked:
                print_entry(sheet_name, entry)
            else:
                print_entry_full(sheet_name, entry)


# ─────────────────────────────────────────────
# 対話モード
# ─────────────────────────────────────────────

def pick_sheet(data: dict) -> str | None:
    names = list(data.keys())
    print("\n📋 シートを選択してください:")
    for i, name in enumerate(names, 1):
        print(f"  {i}. {name}")
    c = input("番号: ").strip()
    try:
        return names[int(c) - 1]
    except (ValueError, IndexError):
        print("❌ 無効な選択です。")
        return None


def interactive_add(filepath: str):
    data = load_all(filepath)
    sheet_name = pick_sheet(data)
    if not sheet_name: return

    is_pin = "PIN" in sheet_name or "暗証番号" in sheet_name

    if is_pin:
        cats = CATEGORIES_PIN
        print("\n👤 所有者を選択:")
        for i, c in enumerate(cats, 1): print(f"  {i}. {c}")
        c = input("番号: ").strip()
        try: cat = cats[int(c)-1]
        except: print("❌ 無効"); return

        entry = {
            "カード/機器名": input("  カード/機器名: ").strip(),
            "PIN/暗証番号":  getpass.getpass("  PIN/暗証番号: "),
            "桁数":         input("  桁数: ").strip(),
            "更新日":       input("  更新日 (例: 2024-01-01): ").strip(),
            "備考":         input("  備考: ").strip(),
        }
        add_entry(filepath, sheet_name, cat, entry)
    else:
        print("\n📂 カテゴリを選択:")
        cats = CATEGORIES_PASSWORD
        for i, c in enumerate(cats, 1): print(f"  {i}. {c}")
        c = input("番号: ").strip()
        try: cat = cats[int(c)-1]
        except: print("❌ 無効"); return

        entry = {
            "サービス名": input("  サービス名: ").strip(),
            "ログインID": input("  ログインID（メール等）: ").strip(),
            "パスワード": getpass.getpass("  パスワード: "),
            "PIN/暗証番号": input("  PIN/暗証番号（任意）: ").strip(),
            "2FA":        input("  2FA設定（例: Google Authenticator / SMS / なし）: ").strip(),
            "バックアップコード": input("  バックアップコード（任意）: ").strip(),
            "備考":       input("  備考（任意）: ").strip(),
        }
        add_entry(filepath, sheet_name, cat, entry)


def interactive_menu(filepath: str):
    print("\n🔐 パスワード管理ツール")
    print("─" * 30)
    print("1. 全エントリを表示（パスワードはマスク）")
    print("2. キーワード検索")
    print("3. エントリを追加")
    print("4. エントリを削除")
    print("0. 終了")
    print("─" * 30)
    choice = input("選択: ").strip()

    if choice == "1":
        data = load_all(filepath)
        list_all(data, masked=True)

    elif choice == "2":
        keyword = input("検索キーワード: ").strip()
        data = load_all(filepath)
        results = search_entries(data, keyword)
        if results:
            print(f"\n🔍 '{keyword}' の検索結果 ({len(results)}件):")
            show_pw = input("パスワードをそのまま表示しますか？ (y/N): ").strip().lower() == "y"
            for sn, entry in results:
                if show_pw:
                    print_entry_full(sn, entry)
                else:
                    print_entry(sn, entry)
        else:
            print(f"⚠️  '{keyword}' は見つかりませんでした。")

    elif choice == "3":
        interactive_add(filepath)

    elif choice == "4":
        data = load_all(filepath)
        sheet_name = pick_sheet(data)
        if not sheet_name: return
        name = input("削除するサービス名/カード名: ").strip()
        if input(f"⚠️  '{name}' を削除します。よろしいですか？ (y/N): ").strip().lower() == "y":
            delete_entry(filepath, sheet_name, name)
        else:
            print("キャンセルしました。")

    elif choice == "0":
        print("終了します。")
        sys.exit(0)
    else:
        print("❌ 無効な選択です。")


# ─────────────────────────────────────────────
# エントリーポイント
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="パスワード管理ファイル操作ツール",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用例:
  python password_vault_manager.py                     # 対話メニュー
  python password_vault_manager.py --list              # 全エントリ一覧（マスク）
  python password_vault_manager.py --list --unmask     # 全エントリ一覧（マスクなし）
  python password_vault_manager.py --search Amazon     # キーワード検索
  python password_vault_manager.py --file my.xlsx      # ファイル指定
        """
    )
    parser.add_argument("--file",    default=VAULT_FILE,  help=f"対象ファイル（デフォルト: {VAULT_FILE}）")
    parser.add_argument("--list",    action="store_true", help="全エントリを表示")
    parser.add_argument("--unmask",  action="store_true", help="パスワードをマスクせず表示")
    parser.add_argument("--search",  metavar="KEYWORD",   help="キーワードで検索")
    args = parser.parse_args()

    if args.list:
        list_all(load_all(args.file), masked=not args.unmask)
    elif args.search:
        results = search_entries(load_all(args.file), args.search)
        if results:
            print(f"\n🔍 '{args.search}' の検索結果 ({len(results)}件):")
            for sn, entry in results:
                print_entry(sn, entry)
        else:
            print(f"⚠️  '{args.search}' は見つかりませんでした。")
    else:
        while True:
            interactive_menu(args.file)


if __name__ == "__main__":
    main()
